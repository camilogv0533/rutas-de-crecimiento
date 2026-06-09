#!/usr/bin/env python3
"""One-time / idempotent: parse the Excel master (Curated Skills + Referentes sheets)
into data/master_queue.json — an ORDERED queue of candidate skill-retreat URLs.

queue_ingest.py pops from this file weekly. The skill-development relevance gate
lives in scraper.py (is_skill_development_retreat), so junk that slips through the
domain pre-filter here is still rejected at scrape time — this file just avoids
spending LLM on obvious non-retreats.

Usage:
  python scripts/build_master_queue.py --xlsx "/Users/camilogv/Downloads/0 Master (1).xlsx"
"""
import argparse
import json
import re
import sqlite3
from pathlib import Path
from urllib.parse import urlparse

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "retreats.db"
SOURCES_PATH = ROOT / "data" / "sources.json"
QUEUE_PATH = ROOT / "data" / "master_queue.json"

URL_RE = re.compile(r"https?://[^\s|,)\"']+")

# Domains that are NOT a single scrapeable skill-retreat page: social, software,
# habit apps, generic search/aggregator landings, news. Skip to save LLM spend.
BLOCK_DOMAINS = {
    "tiktok.com", "instagram.com", "facebook.com", "youtube.com", "twitter.com", "x.com",
    "nomadlist.com", "habitify.me", "habitsgarden.com", "routines.club", "mint.intuit.com",
    "airbnb.com", "airbnb.com.ec", "getapp.com", "bookyway.com", "software.retreat.guru",
    "globalnews.booking.com", "nationalskillsnetwork.in", "skillsyouneed.com",
    "exe-coll.ac.uk", "lepaya.com", "transformational.travel",
}


def host_of(url: str) -> str:
    net = urlparse(url).netloc.lower()
    return net[4:] if net.startswith("www.") else net


def blocked(url: str) -> bool:
    h = host_of(url)
    return any(h == b or h.endswith("." + b) for b in BLOCK_DOMAINS)


def known_urls() -> set:
    known = set()
    if SOURCES_PATH.exists():
        for s in json.loads(SOURCES_PATH.read_text()).get("sources", []):
            known.add(s["url"].rstrip("/"))
    if DB_PATH.exists():
        conn = sqlite3.connect(DB_PATH)
        for (u,) in conn.execute("SELECT source_url FROM retreats"):
            known.add(u.rstrip("/"))
        conn.close()
    return known


def extract(xlsx: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    known = known_urls()
    seen = set()
    out = []

    def add(url: str, skill_hint, pool, sheet):
        url = url.rstrip("/")
        key = url.lower()
        if not url or key in seen or url.rstrip("/") in known or blocked(url):
            return
        seen.add(key)
        out.append({
            "url": url,
            "skill_hint": (skill_hint or "").strip() or None,
            "pool": pool,            # "master" (ordered curated) | "referent" (random pool)
            "source_sheet": sheet,
            "status": "pending",     # pending | added | rejected | failed
            "attempts": 0,
            "added_at": None,
            "note": None,
        })

    # 1) Curated Skills and Experiences — col A = skill name, cols B+ = experience URLs
    if "Curated Skills and Experiences" in wb.sheetnames:
        ws = wb["Curated Skills and Experiences"]
        current_skill = None
        for row in ws.iter_rows(values_only=True):
            cells = [c for c in row if c]
            if not cells:
                continue
            first = str(row[0]).strip() if row[0] else ""
            if first and not first.lower().startswith("http"):
                current_skill = first  # section / skill label
            for c in row:
                if c and isinstance(c, str):
                    for m in URL_RE.findall(c):
                        add(m, current_skill, "master", "curated")

    # 2) Referentes sheets — hosts that run interesting skill retreats (random pool)
    for sheet in ("Referentes ", "Referentes claves", "Transformative travel, communit"):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            comment = None
            # comment column (2nd cell) gives a hint of why it's a referent
            if len(row) > 1 and row[1] and isinstance(row[1], str) and not row[1].lower().startswith("http"):
                comment = row[1]
            for c in row:
                if c and isinstance(c, str):
                    for m in URL_RE.findall(c):
                        add(m, comment, "referent", sheet.strip())
    wb.close()
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="/Users/camilogv/Downloads/0 Master (1).xlsx")
    ap.add_argument("--merge", action="store_true",
                    help="Merge into existing queue (preserve statuses) instead of overwriting.")
    args = ap.parse_args()

    fresh = extract(Path(args.xlsx))

    if args.merge and QUEUE_PATH.exists():
        existing = {e["url"].rstrip("/"): e for e in json.loads(QUEUE_PATH.read_text())}
        for e in fresh:
            existing.setdefault(e["url"].rstrip("/"), e)
        result = list(existing.values())
    else:
        result = fresh

    QUEUE_PATH.write_text(json.dumps(result, indent=2, ensure_ascii=False))
    by_pool = {}
    for e in result:
        by_pool[e["pool"]] = by_pool.get(e["pool"], 0) + 1
    print(f"master_queue.json: {len(result)} candidates {by_pool} → {QUEUE_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
